from io import BytesIO
from openpyxl import Workbook
from backend.application.interfaces.report_generator import ReportGenerator
from backend.application.dtos import ProcessCsvOutput

class ExcelReportGenerator(ReportGenerator):

    def generate_report(self, output_dto: ProcessCsvOutput) -> BytesIO:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary_sheet.cell(row=1, column=1, value="Metric")
        summary_sheet.cell(row=1, column=2, value="Value")

        summary_sheet.cell(row=2, column=1, value="Total Revenue")
        summary_sheet.cell(row=2, column=2, value=output_dto.total_revenue)

        summary_sheet.cell(row=3, column=1, value="MoM Growth (%)")
        summary_sheet.cell(row=3, column=2, value=output_dto.mom_growth_rate)

        summary_sheet.cell(row=4, column=1, value="Avg. Ticket")
        summary_sheet.cell(row=4, column=2, value=output_dto.average_ticket)

        summary_sheet.cell(row=5, column=1, value="New Leads")
        summary_sheet.cell(row=5, column=2, value=output_dto.total_leads)

        monthly_sheet = workbook.create_sheet(title="Monthly Revenue")

        monthly_sheet.cell(row=1, column=1, value="Month")
        monthly_sheet.cell(row=1, column=2, value="Revenue")

        row_idx = 2
        for month, revenue in output_dto.monthly_revenue.items():
            monthly_sheet.cell(row=row_idx, column=1, value=month)
            monthly_sheet.cell(row=row_idx, column=2, value=revenue)
            row_idx += 1

        output_buffer = BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)

        return output_buffer
